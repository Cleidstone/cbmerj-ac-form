import json
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from calculator import (get_calculation_breakdown, get_emop_codes,
                        ROOM_TYPE_FACTORS, format_btu)

# ── Cores ──────────────────────────────────────────────────────────────────
COR_AZUL_ESCURO  = '1B3A6B'
COR_AZUL_MEDIO   = '2E5FA3'
COR_AZUL_CLARO   = 'D6E4F7'
COR_CINZA        = 'F2F2F2'
COR_AMARELO      = 'FFF2CC'
COR_VERDE_CLARO  = 'E2EFDA'
COR_LARANJA      = 'FCE4D6'
BRANCO           = 'FFFFFF'

THIN = Side(style='thin', color='AAAAAA')
MEDIUM = Side(style='medium', color='444444')
BORDA_THIN  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDA_MEDIUM = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)


def _font(bold=False, size=11, color='000000', italic=False):
    return Font(name='Arial', bold=bold, size=size, color=color, italic=italic)


def _fill(hex_color):
    return PatternFill('solid', start_color=hex_color, fgColor=hex_color)


def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def _merge_write(ws, row, col_start, col_end, value, font=None, fill=None,
                 align=None, border=None, number_format=None):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=value)
    if font:          cell.font = font
    if fill:          cell.fill = fill
    if align:         cell.alignment = align
    if border:        cell.border = border
    if number_format: cell.number_format = number_format
    return cell


def _write(ws, row, col, value, font=None, fill=None, align=None, border=None,
           number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:          cell.font = font
    if fill:          cell.fill = fill
    if align:         cell.alignment = align
    if border:        cell.border = border
    if number_format: cell.number_format = number_format
    return cell


# ── Cabeçalho institucional ────────────────────────────────────────────────
def _header(ws, title, sub_data):
    _merge_write(ws, 1, 1, 7,
                 'CORPO DE BOMBEIROS MILITAR DO ESTADO DO RIO DE JANEIRO',
                 font=_font(bold=True, size=13, color=BRANCO),
                 fill=_fill(COR_AZUL_ESCURO),
                 align=_align('center'))
    _merge_write(ws, 2, 1, 7, title,
                 font=_font(bold=True, size=11, color=BRANCO),
                 fill=_fill(COR_AZUL_MEDIO),
                 align=_align('center'))
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    row = 4
    for label, value in sub_data:
        _write(ws, row, 1, label, font=_font(bold=True), fill=_fill(COR_AZUL_CLARO),
               align=_align(), border=BORDA_THIN)
        _merge_write(ws, row, 2, 4, value, font=_font(), fill=_fill(BRANCO),
                     align=_align(), border=BORDA_THIN)
        row += 1
    return row + 1


# ── Planilha: Memória de Cálculo ───────────────────────────────────────────
def _sheet_memoria(wb, sub_data, rooms):
    ws = wb.create_sheet('Memória de Cálculo')
    for c, w in enumerate([4, 30, 16, 16, 16, 14, 18], start=1):
        _set_col_width(ws, c, w)

    row = _header(ws, 'MEMÓRIA DE CÁLCULO DE CARGA TÉRMICA', sub_data)
    row += 1

    # Premissas
    _merge_write(ws, row, 1, 7, 'PREMISSAS DO CÁLCULO (conforme BM4)',
                 font=_font(bold=True, size=10, color=BRANCO),
                 fill=_fill(COR_AZUL_MEDIO), align=_align('center'))
    row += 1
    premissas = [
        '1. Todas as janelas possuem películas de proteção à incidência de raios solares.',
        '2. Todas as lajes possuem proteção de telhado de fibrocimento com inclinação mínima de 9%.',
        '3. As trocas de ar ocorrem através do abrir e fechar das portas e frestas.',
    ]
    for p in premissas:
        _merge_write(ws, row, 1, 7, p, font=_font(size=9, italic=True),
                     fill=_fill(COR_AMARELO), align=_align(wrap=True),
                     border=BORDA_THIN)
        ws.row_dimensions[row].height = 14
        row += 1
    row += 1

    for idx, room in enumerate(rooms, start=1):
        # Título do ambiente
        _merge_write(ws, row, 1, 7,
                     f'AMBIENTE {idx}: {room["name"].upper()}',
                     font=_font(bold=True, size=11, color=BRANCO),
                     fill=_fill(COR_AZUL_ESCURO), align=_align('center'))
        ws.row_dimensions[row].height = 18
        row += 1

        # Dados gerais
        tipo_label = ROOM_TYPE_FACTORS.get(
            room['room_type'], ROOM_TYPE_FACTORS['standard'])['label']
        info = [
            ('Localização / Descrição', room['name'], 'Tipo de Ambiente', tipo_label),
            ('Comprimento (m)', f'{room["length"]:.2f}', 'Largura (m)', f'{room["width"]:.2f}'),
            ('Pé-direito (m)', f'{room["height"]:.2f}', 'Área (m²)', f'{room["area"]:.2f}'),
        ]
        for label1, val1, label2, val2 in info:
            _write(ws, row, 1, label1, font=_font(bold=True), fill=_fill(COR_AZUL_CLARO),
                   align=_align(), border=BORDA_THIN)
            _merge_write(ws, row, 2, 3, val1, font=_font(), fill=_fill(BRANCO),
                         align=_align(), border=BORDA_THIN)
            _write(ws, row, 4, label2, font=_font(bold=True), fill=_fill(COR_AZUL_CLARO),
                   align=_align(), border=BORDA_THIN)
            _merge_write(ws, row, 5, 7, val2, font=_font(), fill=_fill(BRANCO),
                         align=_align(), border=BORDA_THIN)
            row += 1
        row += 1

        # Tabela de cálculo
        headers = ['Variável', 'Quantidade', 'Fator (BTU)', 'Subtotal BTU/h']
        col_spans = [(1, 3), (4, 4), (5, 5), (6, 7)]
        for h, (cs, ce) in zip(headers, col_spans):
            _merge_write(ws, row, cs, ce, h,
                         font=_font(bold=True, color=BRANCO),
                         fill=_fill(COR_AZUL_MEDIO), align=_align('center'),
                         border=BORDA_THIN)
        row += 1

        breakdown = get_calculation_breakdown(
            room['area'], room['people'], room['appliances'],
            room['lamps'], room['window_morning'], room['window_afternoon'],
            room['room_type']
        )
        total_btu = 0
        for var, qty, factor, subtotal in breakdown:
            fill_r = _fill(BRANCO) if row % 2 == 0 else _fill(COR_CINZA)
            _merge_write(ws, row, 1, 3, var, font=_font(), fill=fill_r,
                         align=_align(), border=BORDA_THIN)
            _write(ws, row, 4, qty if isinstance(qty, (int, float)) else str(qty),
                   font=_font(), fill=fill_r, align=_align('center'),
                   border=BORDA_THIN)
            _write(ws, row, 5, factor, font=_font(), fill=fill_r,
                   align=_align('center'), border=BORDA_THIN,
                   number_format='#,##0.00')
            _merge_write(ws, row, 6, 7, subtotal, font=_font(), fill=fill_r,
                         align=_align('right'), border=BORDA_THIN,
                         number_format='#,##0.00')
            total_btu += subtotal
            row += 1

        # Total
        _merge_write(ws, row, 1, 5, 'TOTAL BTU/h CALCULADO',
                     font=_font(bold=True, color=BRANCO),
                     fill=_fill(COR_AZUL_ESCURO), align=_align('right'),
                     border=BORDA_THIN)
        _merge_write(ws, row, 6, 7, round(total_btu, 2),
                     font=_font(bold=True, size=12, color=BRANCO),
                     fill=_fill(COR_AZUL_ESCURO), align=_align('right'),
                     border=BORDA_THIN, number_format='#,##0.00')
        row += 1
        row += 1

        # Equipamento selecionado
        _merge_write(ws, row, 1, 7, 'EQUIPAMENTO DIMENSIONADO',
                     font=_font(bold=True, color=BRANCO),
                     fill=_fill(COR_AZUL_MEDIO), align=_align('center'))
        row += 1

        rec_size = room['recommended_size']
        sel_size = room['selected_size']
        sel_qty  = room['selected_qty']
        is_custom = sel_size not in [12000, 18000, 36000, 60000]
        eq_fill = _fill(COR_LARANJA) if is_custom else _fill(COR_VERDE_CLARO)

        dim_rows = [
            ('Capacidade Recomendada (BTU/h)', f'{rec_size:,} BTU/h'.replace(',', '.')),
            ('Capacidade Selecionada (BTU/h)', f'{sel_size:,} BTU/h'.replace(',', '.')),
            ('Quantidade de Equipamentos', str(sel_qty)),
            ('Capacidade Total Instalada', f'{sel_size * sel_qty:,} BTU/h'.replace(',', '.')),
        ]
        for label, value in dim_rows:
            _write(ws, row, 1, label, font=_font(bold=True),
                   fill=_fill(COR_AZUL_CLARO), align=_align(), border=BORDA_THIN)
            _merge_write(ws, row, 2, 7, value, font=_font(), fill=eq_fill,
                         align=_align(), border=BORDA_THIN)
            row += 1

        if room.get('justification'):
            _write(ws, row, 1, 'Justificativa (fuga ao padrão BM4)',
                   font=_font(bold=True), fill=_fill(COR_LARANJA),
                   align=_align(), border=BORDA_THIN)
            _merge_write(ws, row, 2, 7, room['justification'],
                         font=_font(italic=True), fill=_fill(COR_AMARELO),
                         align=_align(wrap=True), border=BORDA_THIN)
            ws.row_dimensions[row].height = 30
            row += 1

        # Códigos EMOP
        emop = get_emop_codes(sel_size)
        _merge_write(ws, row, 1, 7, 'CÓDIGOS EMOP PARA LICITAÇÃO',
                     font=_font(bold=True, color=BRANCO),
                     fill=_fill(COR_AZUL_MEDIO), align=_align('center'))
        row += 1
        emop_items = [
            (emop['supply_code'],  f'{emop["supply_desc"]} – Qtd: {sel_qty}'),
            (emop['install_code'], f'{emop["install_desc"]} – Qtd: {sel_qty}'),
            (emop['pipe_code'],    f'{emop["pipe_desc"]} – A definir por medição'),
        ]
        for code, desc in emop_items:
            _write(ws, row, 1, code, font=_font(bold=True),
                   fill=_fill(COR_CINZA), align=_align('center'), border=BORDA_THIN)
            _merge_write(ws, row, 2, 7, desc, font=_font(),
                         fill=_fill(BRANCO), align=_align(wrap=True),
                         border=BORDA_THIN)
            row += 1

        row += 2

    return ws


# ── Planilha: Resumo para Licitação ──────────────────────────────────────
def _sheet_resumo(wb, rooms):
    ws = wb.create_sheet('Resumo para Licitação')
    for c, w in enumerate([6, 50, 20, 20], start=1):
        _set_col_width(ws, c, w)

    _merge_write(ws, 1, 1, 4, 'RESUMO DE EQUIPAMENTOS PARA PROCESSO LICITATÓRIO',
                 font=_font(bold=True, size=12, color=BRANCO),
                 fill=_fill(COR_AZUL_ESCURO), align=_align('center'))
    ws.row_dimensions[1].height = 22

    row = 3
    headers = ['EMOP', 'Descrição', 'Qtd.', 'Obs.']
    for c, h in enumerate(headers, start=1):
        _write(ws, row, c, h, font=_font(bold=True, color=BRANCO),
               fill=_fill(COR_AZUL_MEDIO), align=_align('center'),
               border=BORDA_THIN)
    row += 1

    # Contabilizar por tamanho
    from collections import defaultdict
    counts = defaultdict(int)
    for room in rooms:
        sz = room['selected_size']
        qty = room['selected_qty']
        counts[sz] += qty

    supply_emop = {
        12000: '18.030.0002-0', 18000: '18.030.0003-0',
        36000: '18.030.0008-0', 60000: '18.030.0010-0',
    }
    supply_desc = {
        12000: 'Condicionador de ar SPLIT 12.000 BTU/h – Fornecimento',
        18000: 'Condicionador de ar SPLIT 18.000 BTU/h – Fornecimento',
        36000: 'Condicionador de ar SPLIT 36.000 BTU/h – Fornecimento',
        60000: 'Condicionador de ar SPLIT 60.000 BTU/h – Fornecimento',
    }

    small_qty = sum(v for k, v in counts.items() if k <= 30000)
    large_qty = sum(v for k, v in counts.items() if k > 30000)

    for size in [12000, 18000, 36000, 60000]:
        qty = counts.get(size, 0)
        fill_r = _fill(COR_VERDE_CLARO) if qty > 0 else _fill(COR_CINZA)
        _write(ws, row, 1, supply_emop[size], font=_font(bold=bool(qty)),
               fill=fill_r, align=_align('center'), border=BORDA_THIN)
        _write(ws, row, 2, supply_desc[size], font=_font(),
               fill=fill_r, align=_align(), border=BORDA_THIN)
        _write(ws, row, 3, qty if qty else '–', font=_font(bold=bool(qty)),
               fill=fill_r, align=_align('center'), border=BORDA_THIN)
        _write(ws, row, 4, '', font=_font(), fill=fill_r,
               align=_align(), border=BORDA_THIN)
        row += 1

    row += 1
    _merge_write(ws, row, 1, 4, 'SERVIÇOS DE INSTALAÇÃO E TUBULAÇÃO',
                 font=_font(bold=True, color=BRANCO),
                 fill=_fill(COR_AZUL_MEDIO), align=_align('center'))
    row += 1

    service_items = [
        ('15.005.0215-0', 'Assentamento Split 9.000 a 30.000 BTU/h', small_qty, ''),
        ('15.005.0220-0', 'Assentamento Split 36.000 a 60.000 BTU/h', large_qty, ''),
        ('15.005.0240-0', 'Tubulação cobre Split 9.000 a 30.000 BTU/h', small_qty, 'Qtd. de serviço (m) a definir por medição'),
        ('15.005.0245-0', 'Tubulação cobre Split 36.000 a 60.000 BTU/h', large_qty, 'Qtd. de serviço (m) a definir por medição'),
    ]
    for code, desc, qty, obs in service_items:
        fill_r = _fill(COR_VERDE_CLARO) if qty > 0 else _fill(COR_CINZA)
        _write(ws, row, 1, code, font=_font(bold=bool(qty)),
               fill=fill_r, align=_align('center'), border=BORDA_THIN)
        _write(ws, row, 2, desc, font=_font(), fill=fill_r,
               align=_align(), border=BORDA_THIN)
        _write(ws, row, 3, qty if qty else '–', font=_font(bold=bool(qty)),
               fill=fill_r, align=_align('center'), border=BORDA_THIN)
        _write(ws, row, 4, obs, font=_font(italic=True, size=9),
               fill=fill_r, align=_align(wrap=True), border=BORDA_THIN)
        row += 1

    return ws


# ── Planilha: Capa ─────────────────────────────────────────────────────────
def _sheet_capa(wb, sub, rooms):
    ws = wb.active
    ws.title = 'Capa'
    for c, w in enumerate([4, 30, 16, 16, 16, 14, 18], start=1):
        _set_col_width(ws, c, w)

    sub_data = [
        ('Unidade (OBM)', sub['obm_name']),
        ('Responsável', sub['commander_name'] or ''),
        ('E-mail', sub['contact_email'] or ''),
        ('Telefone', sub['contact_phone'] or ''),
        ('Data de Envio', sub['submitted_at']),
    ]
    row = _header(ws, 'MEMÓRIA DE CÁLCULO DE CARGA TÉRMICA – AQUISIÇÃO DE ARS CONDICIONADOS',
                  sub_data)
    row += 1

    # Resumo
    _merge_write(ws, row, 1, 7, 'RESUMO GERAL',
                 font=_font(bold=True, color=BRANCO),
                 fill=_fill(COR_AZUL_ESCURO), align=_align('center'))
    row += 1

    from collections import defaultdict, Counter
    size_counts = Counter()
    for r in rooms:
        size_counts[r['selected_size']] += r['selected_qty']

    total_units = sum(size_counts.values())
    _write(ws, row, 1, 'Total de ambientes', font=_font(bold=True),
           fill=_fill(COR_AZUL_CLARO), align=_align(), border=BORDA_THIN)
    _merge_write(ws, row, 2, 7, len(rooms), font=_font(),
                 fill=_fill(BRANCO), align=_align(), border=BORDA_THIN)
    row += 1
    _write(ws, row, 1, 'Total de equipamentos', font=_font(bold=True),
           fill=_fill(COR_AZUL_CLARO), align=_align(), border=BORDA_THIN)
    _merge_write(ws, row, 2, 7, total_units, font=_font(),
                 fill=_fill(BRANCO), align=_align(), border=BORDA_THIN)
    row += 1
    row += 1

    for size in [12000, 18000, 36000, 60000]:
        qty = size_counts.get(size, 0)
        if qty:
            _write(ws, row, 1, f'Split {size//1000}k BTU/h',
                   font=_font(bold=True), fill=_fill(COR_VERDE_CLARO),
                   align=_align(), border=BORDA_THIN)
            _merge_write(ws, row, 2, 7, f'{qty} unidade(s)',
                         font=_font(), fill=_fill(BRANCO),
                         align=_align(), border=BORDA_THIN)
            row += 1

    # Não padronizados
    custom = [(sz, qt) for sz, qt in size_counts.items()
              if sz not in [12000, 18000, 36000, 60000]]
    if custom:
        row += 1
        _merge_write(ws, row, 1, 7,
                     '⚠ EQUIPAMENTOS FORA DO PADRÃO BM4 (ver justificativas na Memória de Cálculo)',
                     font=_font(bold=True, color='C00000'),
                     fill=_fill(COR_LARANJA), align=_align('center'),
                     border=BORDA_THIN)
        row += 1
        for sz, qt in custom:
            _write(ws, row, 1, f'Split {sz:,} BTU/h'.replace(',', '.'),
                   font=_font(bold=True), fill=_fill(COR_LARANJA),
                   align=_align(), border=BORDA_THIN)
            _merge_write(ws, row, 2, 7, f'{qt} unidade(s)',
                         font=_font(), fill=_fill(BRANCO),
                         align=_align(), border=BORDA_THIN)
            row += 1

    # Observações gerais
    if sub.get('observations'):
        row += 1
        _merge_write(ws, row, 1, 7, 'OBSERVAÇÕES GERAIS',
                     font=_font(bold=True, color=BRANCO),
                     fill=_fill(COR_AZUL_MEDIO), align=_align('center'))
        row += 1
        _merge_write(ws, row, 1, 7, sub['observations'],
                     font=_font(italic=True), fill=_fill(COR_AMARELO),
                     align=_align(wrap=True), border=BORDA_THIN)
        ws.row_dimensions[row].height = max(40, len(sub['observations']) // 6)

    return ws


# ── Gerador principal por unidade ──────────────────────────────────────────
def generate_unit_report(submission):
    rooms = json.loads(submission['rooms_json'])
    wb = Workbook()
    sub_data_header = [
        ('Unidade (OBM)', submission['obm_name']),
        ('Responsável', submission['commander_name'] or ''),
        ('E-mail', submission['contact_email'] or ''),
        ('Telefone', submission['contact_phone'] or ''),
        ('Data de Envio', submission['submitted_at']),
    ]
    _sheet_capa(wb, submission, rooms)
    _sheet_memoria(wb, sub_data_header, rooms)
    _sheet_resumo(wb, rooms)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Gerador consolidado (todas as unidades) ───────────────────────────────
def generate_consolidated_report(submissions):
    from collections import defaultdict, Counter
    wb = Workbook()
    ws = wb.active
    ws.title = 'Consolidado'

    for c, w in enumerate([6, 35, 12, 12, 12, 12, 12, 20], start=1):
        _set_col_width(ws, c, w)

    _merge_write(ws, 1, 1, 8,
                 'CBMERJ – CONSOLIDADO DE NECESSIDADES DE AR CONDICIONADO',
                 font=_font(bold=True, size=13, color=BRANCO),
                 fill=_fill(COR_AZUL_ESCURO), align=_align('center'))
    _merge_write(ws, 2, 1, 8,
                 f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}',
                 font=_font(italic=True, size=9, color=BRANCO),
                 fill=_fill(COR_AZUL_MEDIO), align=_align('center'))
    ws.row_dimensions[1].height = 22

    row = 4
    headers = ['#', 'Unidade (OBM)', 'Ambientes',
               'Split 12k', 'Split 18k', 'Split 36k', 'Split 60k', 'Outros']
    for c, h in enumerate(headers, start=1):
        _write(ws, row, c, h, font=_font(bold=True, color=BRANCO),
               fill=_fill(COR_AZUL_MEDIO), align=_align('center'),
               border=BORDA_THIN)
    row += 1

    totals = Counter()
    custom_rows = []   # lista de (obm_name, room_name, btu_calc, sel_size, sel_qty, justification)

    for idx, sub in enumerate(submissions, start=1):
        rooms = json.loads(sub['rooms_json'])
        size_counts = Counter()
        for r in rooms:
            size_counts[r['selected_size']] += r['selected_qty']
            if r['selected_size'] not in [12000, 18000, 36000, 60000]:
                custom_rows.append((
                    sub['obm_name'],
                    r['name'],
                    r.get('btu_calculated', 0),
                    r['selected_size'],
                    r['selected_qty'],
                    r.get('justification', ''),
                ))

        others = sum(qt for sz, qt in size_counts.items()
                     if sz not in [12000, 18000, 36000, 60000])
        fill_r = _fill(BRANCO) if idx % 2 == 0 else _fill(COR_CINZA)

        vals = [idx, sub['obm_name'], len(rooms),
                size_counts.get(12000, ''), size_counts.get(18000, ''),
                size_counts.get(36000, ''), size_counts.get(60000, ''),
                others or '']
        for c, v in enumerate(vals, start=1):
            _write(ws, row, c, v, font=_font(), fill=fill_r,
                   align=_align('center' if c != 2 else 'left'),
                   border=BORDA_THIN)
        for sz in [12000, 18000, 36000, 60000]:
            totals[sz] += size_counts.get(sz, 0)
        row += 1

    # Linha de totais
    row += 1
    _write(ws, row, 1, '', font=_font(), fill=_fill(BRANCO), border=BORDA_THIN)
    _write(ws, row, 2, 'TOTAL GERAL', font=_font(bold=True),
           fill=_fill(COR_AZUL_CLARO), align=_align(), border=BORDA_THIN)
    _write(ws, row, 3, '', font=_font(), fill=_fill(COR_AZUL_CLARO), border=BORDA_THIN)
    for c, sz in enumerate([12000, 18000, 36000, 60000], start=4):
        _write(ws, row, c, totals.get(sz, 0),
               font=_font(bold=True), fill=_fill(COR_AZUL_CLARO),
               align=_align('center'), border=BORDA_THIN)
    _write(ws, row, 8, '', font=_font(), fill=_fill(COR_AZUL_CLARO), border=BORDA_THIN)

    # ── Aba de equipamentos não padronizados ──────────────────────────────
    ws2 = wb.create_sheet('Equipamentos Não Padronizados')
    for c, w in enumerate([6, 35, 30, 18, 18, 10, 50], start=1):
        _set_col_width(ws2, c, w)

    _merge_write(ws2, 1, 1, 7,
                 'CBMERJ – EQUIPAMENTOS FORA DO PADRÃO BM4 (12k / 18k / 36k / 60k BTU/h)',
                 font=_font(bold=True, size=12, color=BRANCO),
                 fill=_fill(COR_AZUL_ESCURO), align=_align('center'))
    _merge_write(ws2, 2, 1, 7,
                 f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}',
                 font=_font(italic=True, size=9, color=BRANCO),
                 fill=_fill(COR_AZUL_MEDIO), align=_align('center'))
    ws2.row_dimensions[1].height = 22

    hrow = 4
    h2_headers = ['#', 'Unidade (OBM)', 'Ambiente', 'BTU Calculado',
                  'BTU Solicitado', 'Qtd.', 'Justificativa']
    for c, h in enumerate(h2_headers, start=1):
        _write(ws2, hrow, c, h, font=_font(bold=True, color=BRANCO),
               fill=_fill(COR_AZUL_MEDIO), align=_align('center'),
               border=BORDA_THIN)
    hrow += 1

    if custom_rows:
        for i, (obm, room_name, btu_calc, sel_size, sel_qty, just) in enumerate(custom_rows, start=1):
            fill_r = _fill(COR_LARANJA) if i % 2 != 0 else _fill(COR_AMARELO)
            row_vals = [i, obm, room_name,
                        f'{round(btu_calc):,} BTU/h'.replace(',', '.'),
                        f'{sel_size:,} BTU/h'.replace(',', '.'),
                        sel_qty,
                        just or '(sem justificativa)']
            aligns = ['center', 'left', 'left', 'center', 'center', 'center', 'left']
            for c, (v, al) in enumerate(zip(row_vals, aligns), start=1):
                _write(ws2, hrow, c, v, font=_font(), fill=fill_r,
                       align=_align(al, wrap=(c == 7)), border=BORDA_THIN)
            ws2.row_dimensions[hrow].height = 30 if just else 15
            hrow += 1
    else:
        _merge_write(ws2, hrow, 1, 7,
                     'Nenhuma unidade solicitou equipamento fora do padrão BM4.',
                     font=_font(italic=True), fill=_fill(COR_CINZA),
                     align=_align('center'), border=BORDA_THIN)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
